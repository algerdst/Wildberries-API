import sys
import time

import requests
import json
from datetime import datetime, timedelta
import os
import glob
import openpyxl

with open('api ключ.txt', 'r', encoding='utf-8-sig') as file:
    for i in file:
        api_key = i

headers = {
    "Authorization": api_key
}



def get_companies_ids_9(headers):
    """
    возвращает список с айдишниками кампаний
    :return:
    """
    url = 'https://advert-api.wildberries.ru/adv/v1/promotion/count'
    response = requests.get(url, headers=headers).json()
    adverts_list = []
    ids = []
    for i in response['adverts']:
        if i['status'] == 9:
            adverts_list = i['advert_list']
    for i in adverts_list:
        id = i['advertId']
        ids.append(id)
    return ids


def get_companies_ids_11(headers):
    """
    возвращает список с айдишниками кампаний
    :return:
    """
    url = 'https://advert-api.wildberries.ru/adv/v1/promotion/count'
    response = requests.get(url, headers=headers).json()
    adverts_list = []
    ids = []
    for i in response['adverts']:
        if i['status'] == 11:
            adverts_list = i['advert_list']
    for i in adverts_list:
        id = i['advertId']
        ids.append(id)
    return ids


def get_keywords_statistics(ids):
    print('СБОР СТАТИСТИКИ ПО КЛЮЧЕВЫМ СЛОВАМ')
    file = []
    path = os.getcwd()
    for filename in glob.glob(os.path.join(path, '*.xlsx')):
        file.append(filename)
    filename = file[0]
    book = openpyxl.load_workbook(filename)
    keyword_statistics_sheet = book.worksheets[3]
    headers = {
        "Authorization": 'eyJhbGciOiJFUzI1NiIsImtpZCI6IjIwMjQwNTA2djEiLCJ0eXAiOiJKV1QifQ.eyJlbnQiOjEsImV4cCI6MTczMzUxMzk4MSwiaWQiOiI3YjE0YzRjMC0yODViLTQzZTUtOGFjZS04YjUxNTYzMTZhMTciLCJpaWQiOjUwODE0NjQyLCJvaWQiOjE1ODk5NCwicyI6NDA5NCwic2lkIjoiYjFhMDdlNWMtZGIzYi00MjQ0LThjYzctYjkzYTJmMTgxMmRjIiwidCI6ZmFsc2UsInVpZCI6NTA4MTQ2NDJ9.rK-CZrfOob8bWaDsVb4UTlsa9VufF3qSeNYfeQjAn5tZkowmtIVkcrESynj9DGmoGI71uqtIOrYKDspLY0J57g'}

    """
    собирает статистику по ключевым словам
    """
    dates=[]
    with open('даты для сбора статистики по ключевым словам.txt', 'r', encoding='utf-8-sig') as file:
        for i in file:
            dates.append(i.replace('\n',''))

    row = 2
    for start_date in dates:
        start_date = datetime.strptime(start_date, "%Y-%m-%d")
        for id in ids:
            url = 'https://advert-api.wildberries.ru/adv/v2/auto/daily-words'
            params = {
                'id': id
            }
            response = requests.get(url, headers=headers, params=params).json()
            for day in response:
                date=datetime.strptime(day['date'][:10], "%Y-%m-%d")
                if date==start_date:
                    print(f"СБОР СТАТИСТИКИ ПО КЛЮЧЕВЫМ СЛОВАМ ДЛЯ КАМПАНИИ - {id} ДАТА - {day['date'][:10]} ")
                    stat=day['stat']
                    for dictionary in stat:
                        keyword=dictionary['keyword']
                        views=dictionary['views']
                        clicks=dictionary['clicks']
                        ctr=dictionary['ctr']
                        sum=dictionary['sum']
                        date=day['date'][:10]
                        keyword_statistics_sheet.cell(column=1, row=row).value = id
                        keyword_statistics_sheet.cell(column=2, row=row).value = keyword
                        keyword_statistics_sheet.cell(column=3, row=row).value = views
                        keyword_statistics_sheet.cell(column=4, row=row).value = clicks
                        keyword_statistics_sheet.cell(column=5, row=row).value = ctr
                        keyword_statistics_sheet.cell(column=6, row=row).value = sum
                        keyword_statistics_sheet.cell(column=7, row=row).value = date
                        row+=1
                    book.save(filename)
                else:
                    continue

ids_9 = get_companies_ids_9(
    headers=headers)  # собирает id кампаний со статусом 9, передает полученный список в функцию get_companies_items_common_info
ids_11 = get_companies_ids_11(
    headers=headers)  # собирает id кампаний со статусом 11, передает полученный список в функцию get_companies_items_common_info
ids=ids_9+ids_11
get_keywords_statistics(ids)