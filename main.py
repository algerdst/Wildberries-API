import sys
import time

import requests
import json
from datetime import datetime, timedelta
import os
import glob
import openpyxl

with open('api ключ.txt', 'r', encoding='utf-8-sig') as file:
    for i in file:
        api_key = i

headers = {
    "Authorization": api_key
}

with open('период получения статистики.txt', 'r', encoding='utf-8-sig') as file:
    for i in file:
        n = int(i)

now = datetime.now()-timedelta(2)
period = timedelta(-n)
before_n_days = str(now + period)[:10]
now = str(now)[:10]

companies_statuses = {
    7: 'Кампания завершена',
    9: 'идут показы',
    11: 'Кампания на паузе'
}

companies_types = {
    4: 'кампания в каталоге',
    5: 'кампания в карточке товара',
    6: 'кампания в поиске',
    7: 'кампания в рекомендациях на главной странице',
    8: 'автоматическая кампания',
    9: 'поиск + каталог'
}


def get_companies_ids_9(headers):
    """
    возвращает список с айдишниками кампаний
    :return:
    """
    url = 'https://advert-api.wildberries.ru/adv/v1/promotion/count'
    response = requests.get(url, headers=headers).json()
    adverts_list = []
    ids = []
    for i in response['adverts']:
        if i['status'] == 9:
            adverts_list = i['advert_list']
    for i in adverts_list:
        id = i['advertId']
        ids.append(id)
    return ids


def get_companies_ids_11(headers):
    """
    возвращает список с айдишниками кампаний
    :return:
    """
    url = 'https://advert-api.wildberries.ru/adv/v1/promotion/count'
    response = requests.get(url, headers=headers).json()
    adverts_list = []
    ids = []
    for i in response['adverts']:
        if i['status'] == 11:
            adverts_list = i['advert_list']
    for i in adverts_list:
        id = i['advertId']
        ids.append(id)
    return ids

def get_companies_items_common_info(headers):
    """
    собирает общие данные о кампаниях, собирает их в словарь
    :param headers:
    :return:
    """
    compamies_common_info={}
    url = f'https://advert-api.wildberries.ru/adv/v1/promotion/adverts'
    response = requests.post(url, headers=headers, data=json.dumps(ids)).json()
    for companie in response:
        compamies_common_info[companie['advertId']]={'status': companies_statuses[companie['status']], 'type': companies_types[companie['type']]}
    return compamies_common_info


def get_statistics(companies_common_info,headers):
    print('СБОР ПОЛНОЙ СТАТИСТИКИ')

    """
    собирает статистику
    :param headers:
    :return:
    """
    file = []
    path = os.getcwd()
    for filename in glob.glob(os.path.join(path, '*.xlsx')):
        file.append(filename)
    filename = file[0]
    book = openpyxl.load_workbook(filename)
    common_statistics_sheet = book.worksheets[0]
    full_statistics_sheet = book.worksheets[1]
    companies_ids_and_intervals = [{"id": company_id, "interval": {"begin": f"{before_n_days}", "end": f"{now}"}} for company_id in
                     ids]
    payload = json.dumps(
        companies_ids_and_intervals
    )
    url = f'https://advert-api.wildberries.ru/adv/v2/fullstats'
    response = requests.post(url, headers=headers, data=payload)
    if response.status_code!=200:
        print(response.text)
        print('попробуйте еще раз')
        time.sleep(10)
        sys.exit()
    # with open('rez.json', 'w', encoding='utf-8') as file:
    #     json.dump(response.json(), file, indent=4, ensure_ascii=False)
    # with open('rez.json', 'r', encoding='utf-8') as file:
    #     rez = json.load(file)
    full_statistics = {}  # общая статистика для листа "Общая статистика"
    for company_id in companies_common_info: #переносит общую статистику из company common info в full statistics
        for company in response.json():
            if company_id == company['advertId']:
                full_statistics[company_id] = {'status': companies_common_info[company_id]['status'],
                                                 'type': companies_common_info[company_id]['type']}
            else:
                continue
    for i in response.json():
        for company_id in full_statistics:
            if i['advertId']==company_id:
                full_statistics[company_id]['views']=i['views']
                full_statistics[company_id]['clicks']=i['clicks']
                full_statistics[company_id]['ctr']=i['ctr']
                full_statistics[company_id]['cpc']=i['cpc']
                full_statistics[company_id]['cr']=i['cr']
                full_statistics[company_id]['atbs']=i['atbs']
                full_statistics[company_id]['orders']=i['orders']
                full_statistics[company_id]['sum']=i['sum']
                assortment_statistics={} # сюда собирается детальная статистика по каждому артикулу из кампании
                days=i['days']
                for day in days:
                    apps=day['apps']
                    for app in apps:
                        nomenclature=app['nm']
                        for n in nomenclature:
                            if n['nmId'] not in assortment_statistics:
                                assortment_statistics[n['nmId']]={}
                            for key in n:
                                if key=='nmId':
                                    continue
                                elif key not in assortment_statistics[n['nmId']]:
                                    assortment_statistics[n['nmId']][key]=n[key]
                                elif key in assortment_statistics[n['nmId']] and key != 'name':
                                    assortment_statistics[n['nmId']][key]+=n[key]
                full_statistics[company_id]['assortment_statistics'] =assortment_statistics
    row=2
    for company_id in full_statistics:
        common_statistics_sheet.cell(column=1, row=row).value = company_id
        common_statistics_sheet.cell(column=2, row=row).value = full_statistics[company_id]['type']
        common_statistics_sheet.cell(column=3, row=row).value = full_statistics[company_id]['status']
        common_statistics_sheet.cell(column=4, row=row).value = full_statistics[company_id]['views']
        common_statistics_sheet.cell(column=5, row=row).value = full_statistics[company_id]['clicks']
        common_statistics_sheet.cell(column=6, row=row).value = full_statistics[company_id]['ctr']
        common_statistics_sheet.cell(column=7, row=row).value = full_statistics[company_id]['cpc']
        common_statistics_sheet.cell(column=8, row=row).value = full_statistics[company_id]['cr']
        common_statistics_sheet.cell(column=9, row=row).value = full_statistics[company_id]['atbs']
        common_statistics_sheet.cell(column=10, row=row).value = full_statistics[company_id]['orders']
        common_statistics_sheet.cell(column=11, row=row).value = full_statistics[company_id]['sum']
        book.save(filename)
        row+=1

    row = 2
    for company_id in full_statistics:
        for nm in full_statistics[company_id]['assortment_statistics']:
            name=full_statistics[company_id]['assortment_statistics'][nm]['name']
            views=full_statistics[company_id]['assortment_statistics'][nm]['views']
            clicks=full_statistics[company_id]['assortment_statistics'][nm]['clicks']
            ctr=full_statistics[company_id]['assortment_statistics'][nm]['ctr']
            cpc=full_statistics[company_id]['assortment_statistics'][nm]['cpc']
            sum=full_statistics[company_id]['assortment_statistics'][nm]['sum']
            atbs=full_statistics[company_id]['assortment_statistics'][nm]['atbs']
            orders=full_statistics[company_id]['assortment_statistics'][nm]['orders']
            cr=full_statistics[company_id]['assortment_statistics'][nm]['cr']
            shks=full_statistics[company_id]['assortment_statistics'][nm]['shks']
            sum_price=full_statistics[company_id]['assortment_statistics'][nm]['sum_price']
            try:
                cpm = full_statistics[company_id]['assortment_statistics'][nm]['cpm']
            except:
                cpm = '-'
            full_statistics_sheet.cell(column=1, row=row).value = company_id
            full_statistics_sheet.cell(column=2, row=row).value = name
            full_statistics_sheet.cell(column=3, row=row).value = nm
            full_statistics_sheet.cell(column=4, row=row).value = views
            full_statistics_sheet.cell(column=5, row=row).value = clicks
            full_statistics_sheet.cell(column=6, row=row).value = ctr
            full_statistics_sheet.cell(column=7, row=row).value = cpm
            full_statistics_sheet.cell(column=8, row=row).value = atbs
            full_statistics_sheet.cell(column=9, row=row).value = cpc
            full_statistics_sheet.cell(column=10, row=row).value = orders
            full_statistics_sheet.cell(column=11, row=row).value = sum_price
            full_statistics_sheet.cell(column=12, row=row).value = sum
            full_statistics_sheet.cell(column=13, row=row).value = cr
            book.save(filename)
            row += 1






ids_9 = get_companies_ids_9(
    headers=headers)  # собирает id кампаний со статусом 9, передает полученный список в функцию get_companies_items_common_info
ids_11 = get_companies_ids_11(
    headers=headers)  # собирает id кампаний со статусом 11, передает полученный список в функцию get_companies_items_common_info
ids=ids_9+ids_11
companies_common_info=get_companies_items_common_info(
    headers=headers)  # собирает общую информацию о кампаниях, формирует словарь, передает его в функцию get_statistics
get_statistics(companies_common_info, headers=headers) #собирает полную информацию о кампаниях, формирует словарь








#эта функция скопирована в другой файл и вызывается там, здесь она на всякий случай
# def get_keywords_statistics(ids):
#     print('СБОР СТАТИСТИКИ ПО КЛЮЧЕВЫМ СЛОВАМ')
#     file = []
#     path = os.getcwd()
#     for filename in glob.glob(os.path.join(path, '*.xlsx')):
#         file.append(filename)
#     filename = file[0]
#     book = openpyxl.load_workbook(filename)
#     keyword_statistics_sheet = book.worksheets[3]
#     headers = {
#         "Authorization": 'eyJhbGciOiJFUzI1NiIsImtpZCI6IjIwMjQwNTA2djEiLCJ0eXAiOiJKV1QifQ.eyJlbnQiOjEsImV4cCI6MTczMzUxMzk4MSwiaWQiOiI3YjE0YzRjMC0yODViLTQzZTUtOGFjZS04YjUxNTYzMTZhMTciLCJpaWQiOjUwODE0NjQyLCJvaWQiOjE1ODk5NCwicyI6NDA5NCwic2lkIjoiYjFhMDdlNWMtZGIzYi00MjQ0LThjYzctYjkzYTJmMTgxMmRjIiwidCI6ZmFsc2UsInVpZCI6NTA4MTQ2NDJ9.rK-CZrfOob8bWaDsVb4UTlsa9VufF3qSeNYfeQjAn5tZkowmtIVkcrESynj9DGmoGI71uqtIOrYKDspLY0J57g'}
#
#     """
#     собирает статистику по ключевым словам
#     """
#     with open('даты для сбора статистики по ключевым словам.txt', 'r', encoding='utf-8-sig') as file:
#         for i in file:
#             start_date = i
#     start_date = datetime.strptime(start_date, "%Y-%m-%d")
#     for id in ids:
#         url = 'https://advert-api.wildberries.ru/adv/v2/auto/daily-words'
#         params = {
#             'id': id
#         }
#         response = requests.get(url, headers=headers, params=params).json()
#         row=2
#         for day in response:
#             Flag=False
#             date=datetime.strptime(day['date'][:10], "%Y-%m-%d")
#             result_days_difference=str(start_date-date)
#             if '-' in result_days_difference or 'd' not in result_days_difference:
#                 print(f"СБОР СТАТИСТИКИ ПО КЛЮЧЕВЫМ СЛОВАМ ДЛЯ КАМПАНИИ - {id} ДАТА - {day['date'][:10]} ")
#                 Flag=True
#                 if Flag:
#                     stat=day['stat']
#                     for dictionary in stat:
#                         keyword=dictionary['keyword']
#                         views=dictionary['views']
#                         clicks=dictionary['clicks']
#                         ctr=dictionary['ctr']
#                         sum=dictionary['sum']
#                         date=day['date'][:10]
#                         keyword_statistics_sheet.cell(column=1, row=row).value = id
#                         keyword_statistics_sheet.cell(column=2, row=row).value = keyword
#                         keyword_statistics_sheet.cell(column=3, row=row).value = views
#                         keyword_statistics_sheet.cell(column=4, row=row).value = clicks
#                         keyword_statistics_sheet.cell(column=5, row=row).value = ctr
#                         keyword_statistics_sheet.cell(column=6, row=row).value = sum
#                         keyword_statistics_sheet.cell(column=7, row=row).value = date
#                         row+=1
#                         book.save(filename)
#             else:
#                 break
# get_keywords_statistics(ids)